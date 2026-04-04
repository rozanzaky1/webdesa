import re
from pathlib import Path
from openpyxl import load_workbook

p = Path(r"D:\Downloads\BISMILLAH SEMPRO FEBRUARI\data_bersih.xlsx")
wb = load_workbook(p, read_only=True, data_only=True)
ws = wb.active

header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
nik_idx = list(header).index('NIK')

NIK_RE = re.compile(r'^\d{16}$')

non_empty = 0
valid = 0
examples_invalid = []

for row in ws.iter_rows(min_row=2, values_only=True):
    v = row[nik_idx]
    if v is None:
        continue
    s = str(v).strip()
    if s.lower() in {'', 'nan', 'none', 'null'}:
        continue
    non_empty += 1

    m = re.match(r'^(\d+)\.0$', s)
    if m:
        s = m.group(1)
    s = re.sub(r'\D+', '', s)

    if NIK_RE.match(s):
        valid += 1
    else:
        if len(examples_invalid) < 10:
            examples_invalid.append((str(v), s, len(s)))

print('Rows with NIK non-empty:', non_empty)
print('Rows with NIK valid 16 digits:', valid)
print('Invalid examples (raw, normalized, len):')
for ex in examples_invalid:
    print(' -', ex)
