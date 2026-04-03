import json
from pathlib import Path
from openpyxl import load_workbook

SOURCE = Path(r"D:\Downloads\BISMILLAH SEMPRO FEBRUARI\data_bersih.xlsx")
HAMLETS_FILE = Path(r"c:\laragon\www\webdesa\storage\hamlets.json")

# Load Excel file
wb = load_workbook(SOURCE, read_only=True, data_only=True)
ws = wb.active

# Get column index untuk Dusun
header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
dusun_idx = None
for i, col in enumerate(header):
    if col and 'dusun' in str(col).lower():
        dusun_idx = i
        break

if dusun_idx is None:
    print("❌ Kolom 'Dusun' tidak ditemukan!")
    exit(1)

# Extract unique hamlets
hamlets_set = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    value = row[dusun_idx]
    if value and str(value).strip().lower() not in {'', 'nan', 'none', 'null'}:
        hamlet_name = str(value).strip()
        hamlets_set.add(hamlet_name)

# Convert to sorted list with IDs
hamlets_list = sorted(list(hamlets_set))
hamlets_data = [{"id": i+1, "name": name} for i, name in enumerate(hamlets_list)]

# Create storage folder
HAMLETS_FILE.parent.mkdir(parents=True, exist_ok=True)

# Write to JSON
with HAMLETS_FILE.open('w', encoding='utf-8') as f:
    json.dump(hamlets_data, f, ensure_ascii=False, indent=2)

print(f"✅ Hamlets file created: {HAMLETS_FILE}")
print(f"Total unique hamlets: {len(hamlets_data)}\n")
print("Hamlets list:")
for h in hamlets_data:
    print(f"  {h['id']}. {h['name']}")
