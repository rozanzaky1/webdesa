from openpyxl import load_workbook
from pathlib import Path

p = Path(r"D:\Downloads\BISMILLAH SEMPRO FEBRUARI\data_bersih.xlsx")
wb = load_workbook(p, read_only=True, data_only=True)
ws = wb.active

header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
nik_idx = list(header).index('NIK')

found = 0
for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    v = row[nik_idx]
    if v is None:
        continue
    s = str(v).strip()
    if s.lower() in {'', 'nan', 'none', 'null'}:
        continue
    print('Row', r, 'value=', v, 'type=', type(v).__name__)
    found += 1
    if found >= 20:
        break

print('Found:', found)
