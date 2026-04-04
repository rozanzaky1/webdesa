from openpyxl import load_workbook
from pathlib import Path

p = Path(r"D:\Downloads\BISMILLAH SEMPRO FEBRUARI\data_bersih.xlsx")
wb = load_workbook(p, read_only=True, data_only=True)
ws = wb.active

header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
first = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))

print('File:', p)
print('Sheet:', ws.title)
print('Columns:', len(header))
print('Header:', header)
print('FirstRow:', first)
print('TotalRows(approx):', ws.max_row)
