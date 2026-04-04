from __future__ import annotations

import csv
import re
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path(r"D:\Downloads\BISMILLAH SEMPRO FEBRUARI\data_bersih.xlsx")
OUT = Path(r"c:\laragon\www\webdesa\storage\app\import\data_bersih_clean.csv")

HEADER_MAP = {
    "nomor kartu keluarga (kk)": "family_card_number",
    "nik": "nik",
    "nama lengkap": "name",
    "jenis kelamin": "gender",
    "tempat lahir": "birth_place",
    "tanggal lahir": "birth_date",
    "alamat": "address",
    "dusun": "hamlet",
    "status perkawinan": "marital_status",
    "pekerjaan": "occupation",
    "status penduduk": "status",
    "agama": "religion",
    "no telepon": "phone",
}

OUT_HEADERS = [
    "nik",
    "family_card_number",
    "name",
    "gender",
    "birth_place",
    "birth_date",
    "address",
    "hamlet",
    "religion",
    "marital_status",
    "occupation",
    "phone",
    "status",
]

NIK_RE = re.compile(r"^\d{16}$")


def _norm(s: object) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    if s.lower() in {"nan", "none", "null"}:
        return ""
    return s


def _digits_only(s: str) -> str:
    s = s.strip()
    # Many numeric identifiers in the sheet are stored like "1802052204920002.0".
    m = re.match(r"^(\d+)\.0$", s)
    if m:
        return m.group(1)
    return re.sub(r"\D+", "", s)


def _parse_gender(raw: str) -> str:
    v = raw.strip().lower()
    if v in {"l", "lk", "laki", "laki-laki", "pria", "male"}:
        return "Male"
    if v in {"p", "pr", "perempuan", "wanita", "female"}:
        return "Female"
    return raw


def _parse_marital(raw: str) -> str:
    v = raw.strip().lower()
    if v in {"belum menikah", "single", "tidak kawin"}:
        return "Single"
    if v in {"sudah menikah", "menikah", "married", "kawin"}:
        return "Married"
    if v in {"pernah menikah", "cerai", "duda", "janda", "divorced", "widowed"}:
        return "Divorced"
    return raw


def _parse_status(raw: str) -> str:
    v = raw.strip().lower()
    if v in {"aktif", "active"}:
        return "active"
    if v in {"pindah", "moved"}:
        return "moved"
    if v in {"meninggal", "deceased"}:
        return "deceased"
    # default to active if empty
    return raw or "active"


def _parse_date(cell: object) -> str:
    if cell is None:
        return ""

    if isinstance(cell, (datetime, date)):
        return cell.strftime("%Y-%m-%d")

    s = _norm(cell)
    if not s:
        return ""

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass

    # Give up; keep original
    return s


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"Source file not found: {SOURCE}")

    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb.active

    raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_keys = [_norm(h).lower() for h in raw_headers]

    col_idx = {}
    for i, key in enumerate(header_keys):
        if key in HEADER_MAP:
            col_idx[HEADER_MAP[key]] = i

    missing = [k for k in ("nik",) if k not in col_idx]
    if missing:
        raise SystemExit(f"Missing required columns in Excel header: {missing}\nHeader was: {raw_headers}")

    OUT.parent.mkdir(parents=True, exist_ok=True)

    total = 0
    written = 0
    skipped_invalid_nik = 0

    with OUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OUT_HEADERS)
        writer.writeheader()

        for row in ws.iter_rows(min_row=2, values_only=True):
            total += 1

            nik_raw = _digits_only(_norm(row[col_idx["nik"]]))
            if not NIK_RE.match(nik_raw):
                skipped_invalid_nik += 1
                continue

            def get(col: str) -> str:
                if col not in col_idx:
                    return ""
                return _norm(row[col_idx[col]])

            record = {
                "nik": nik_raw,
                "family_card_number": _digits_only(get("family_card_number")),
                "name": get("name"),
                "gender": _parse_gender(get("gender")),
                "birth_place": get("birth_place"),
                "birth_date": _parse_date(row[col_idx["birth_date"]]) if "birth_date" in col_idx else "",
                "address": get("address"),
                "hamlet": get("hamlet"),
                "religion": get("religion"),
                "marital_status": _parse_marital(get("marital_status")),
                "occupation": get("occupation"),
                "phone": _digits_only(get("phone")),
                "status": _parse_status(get("status")),
            }

            writer.writerow(record)
            written += 1

    print(f"Source: {SOURCE}")
    print(f"Output: {OUT}")
    print(f"Rows scanned: {total}")
    print(f"Rows written (valid NIK): {written}")
    print(f"Skipped invalid NIK: {skipped_invalid_nik}")


if __name__ == "__main__":
    main()
