import re
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook

SOURCE = Path(r"D:\Downloads\BISMILLAH SEMPRO FEBRUARI\data_bersih.xlsx")
OUT = Path(r"c:\laragon\www\webdesa\storage\app\import\insert_residents.sql")

def _norm(s: object) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    if s.lower() in {"nan", "none", "null"}:
        return ""
    return s

def _digits_only(s: str) -> str:
    s = s.strip()
    m = re.match(r"^(\d+)\.0$", s)
    if m:
        return m.group(1)
    return re.sub(r"\D+", "", s)

def _parse_gender(raw: str) -> str:
    v = raw.strip().lower()
    if v in {"l", "lk", "laki", "laki-laki", "pria", "male"}:
        return "Male"
    if v in {"p", "pr", "perempuan", "wanita", "female"}:
        return "Female"
    return "Male"

def _parse_marital(raw: str) -> str:
    v = raw.strip().lower()
    if v in {"belum menikah", "single", "tidak kawin"}:
        return "Single"
    if v in {"sudah menikah", "menikah", "married", "kawin"}:
        return "Married"
    if v in {"pernah menikah", "cerai", "duda", "janda", "divorced", "widowed"}:
        return "Divorced"
    return "Single"

def _parse_status(raw: str) -> str:
    v = raw.strip().lower()
    if v in {"aktif", "active"}:
        return "active"
    if v in {"pindah", "moved"}:
        return "moved"
    if v in {"meninggal", "deceased"}:
        return "deceased"
    return "active"

def _parse_date(cell: object) -> str:
    if cell is None:
        return "NULL"
    
    if isinstance(cell, (datetime, date)):
        return f"'{cell.strftime('%Y-%m-%d')}'"
    
    s = _norm(cell)
    if not s:
        return "NULL"
    
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            result = datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            return f"'{result}'"
        except ValueError:
            pass
    
    return "NULL"

def _sql_value(v: str) -> str:
    if not v or v.lower() in {"", "null"}:
        return "NULL"
    v = v.replace("'", "\\'")
    return f"'{v}'"

wb = load_workbook(SOURCE, read_only=True, data_only=True)
ws = wb.active

raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
header_keys = [_norm(h).lower() for h in raw_headers]

HEADER_MAP = {
    "nomor kartu keluarga (kk)": "family_card_number",
    "nik": "nik",
    "nama lengkap": "name",
    "jenis kelamin": "gender",
    "tempat lahir": "birth_place",
    "tanggal lahir": "birth_date",
    "alamat": "address",
    "dusun": "hamlet",
    "status perkawinan": "marital_status",
    "pekerjaan": "occupation",
    "status penduduk": "status",
    "agama": "religion",
    "no telepon": "phone",
}

col_idx = {}
for i, key in enumerate(header_keys):
    if key in HEADER_MAP:
        col_idx[HEADER_MAP[key]] = i

NIK_RE = re.compile(r"^\d{16}$")

records = []
skipped = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    nik_raw = _digits_only(_norm(row[col_idx["nik"]]))
    if not NIK_RE.match(nik_raw):
        skipped += 1
        continue
    
    def get(col: str) -> str:
        if col not in col_idx:
            return ""
        return _norm(row[col_idx[col]])
    
    record = {
        "nik": nik_raw,
        "family_card_number": _digits_only(get("family_card_number")),
        "name": get("name"),
        "gender": _parse_gender(get("gender")),
        "birth_place": get("birth_place"),
        "birth_date": _parse_date(row[col_idx["birth_date"]]) if "birth_date" in col_idx else "NULL",
        "address": get("address"),
        "hamlet": get("hamlet"),
        "religion": get("religion"),
        "marital_status": _parse_marital(get("marital_status")),
        "occupation": get("occupation"),
        "phone": _digits_only(get("phone")),
        "status": _parse_status(get("status")),
    }
    records.append(record)

with OUT.open("w", encoding="utf-8") as f:
    f.write("-- Import data penduduk dari Excel\n")
    f.write("-- Auto-generated SQL INSERT script\n")
    f.write(f"-- Total records: {len(records)}\n")
    f.write("-- Database: laravel_desa\n")
    f.write("-- Table: residents\n\n")
    f.write("-- Option 1: Disable foreign key checks (safer for large imports)\n")
    f.write("SET FOREIGN_KEY_CHECKS=0;\n\n")
    
    f.write("INSERT INTO residents (nik, family_card_number, name, gender, birth_place, birth_date, address, hamlet, religion, marital_status, occupation, phone, status, created_at, updated_at) VALUES\n")
    
    values = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    for rec in records:
        val = (
            f"{_sql_value(rec['nik'])}, "
            f"{_sql_value(rec['family_card_number'])}, "
            f"{_sql_value(rec['name'])}, "
            f"{_sql_value(rec['gender'])}, "
            f"{_sql_value(rec['birth_place'])}, "
            f"{rec['birth_date']}, "
            f"{_sql_value(rec['address'])}, "
            f"{_sql_value(rec['hamlet'])}, "
            f"{_sql_value(rec['religion'])}, "
            f"{_sql_value(rec['marital_status'])}, "
            f"{_sql_value(rec['occupation'])}, "
            f"{_sql_value(rec['phone'])}, "
            f"{_sql_value(rec['status'])}, "
            f"'{now}', '{now}'"
        )
        values.append(f"({val})")
    
    f.write(",\n".join(values))
    f.write(";\n\n")
    
    f.write("-- Re-enable foreign key checks\n")
    f.write("SET FOREIGN_KEY_CHECKS=1;\n\n")
    
    f.write(f"-- Verification query\n")
    f.write(f"SELECT COUNT(*) as total_inserted FROM residents WHERE nik IS NOT NULL;\n")

print(f"SQL file generated: {OUT}")
print(f"Total records: {len(records)}")
print(f"Skipped invalid NIK: {skipped}")
print(f"File size: {OUT.stat().st_size / 1024:.1f} KB")
